from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from app.config import Settings


@dataclass(frozen=True)
class CachedTemplate:
    marketplace: str
    category_id: int
    type_id: int | None
    path: Path


def template_path(settings: Settings, marketplace: str, category_id: int, type_id: int | None = None) -> Path:
    suffix = f"_{type_id}" if marketplace == "ozon" and type_id else ""
    return settings.TEMPLATE_CACHE_DIR / marketplace / f"{category_id}{suffix}" / "template.xlsx"


def find_template(settings: Settings, marketplace: str, category_id: int, type_id: int | None = None) -> CachedTemplate | None:
    direct = template_path(settings, marketplace, category_id, type_id)
    fallback = template_path(settings, marketplace, category_id, None)
    for path in (direct, fallback):
        if path.exists():
            return CachedTemplate(marketplace, category_id, type_id, path)
    return None


def save_template_bytes(
    settings: Settings,
    *,
    content: bytes,
    marketplace: str | None = None,
    category_id: int | None = None,
    type_id: int | None = None,
) -> CachedTemplate:
    if marketplace is None or category_id is None:
        detected_marketplace, detected_category_id, detected_type_id = infer_template_identity(content)
        marketplace = marketplace or detected_marketplace
        category_id = category_id or detected_category_id
        type_id = type_id or detected_type_id
    if marketplace not in {"ozon", "wb"}:
        raise ValueError("template marketplace must be ozon or wb")
    if not category_id:
        raise ValueError("template category_id is required")

    path = template_path(settings, marketplace, category_id, type_id)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(content)
    return CachedTemplate(marketplace, category_id, type_id, path)


def parse_template_hint(text: str) -> tuple[str | None, int | None, int | None]:
    lower = text.lower()
    marketplace = "ozon" if "ozon" in lower or "озон" in lower else "wb" if "wb" in lower or "wildberries" in lower or "вб" in lower else None
    numbers = [int(item) for item in re.findall(r"\b\d{3,}\b", text)]
    category_id = numbers[0] if numbers else None
    type_id = numbers[1] if marketplace == "ozon" and len(numbers) > 1 else None
    return marketplace, category_id, type_id


def infer_template_identity(content: bytes) -> tuple[str | None, int | None, int | None]:
    import io

    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheets = set(workbook.sheetnames)
    if "configs" in sheets:
        ws = workbook["configs"]
        category_id: int | None = None
        type_id: int | None = None
        for row in ws.iter_rows(min_row=1, max_row=30, max_col=2, values_only=True):
            key = str(row[0] or "").strip().upper()
            value = row[1]
            if key == "DESCRIPTION_CATEGORY_ID":
                category_id = _int_or_none(value)
            elif key in {"TYPE_ID", "PRODUCTS_TYPE_ID"}:
                type_id = _int_or_none(value)
        if category_id:
            return "ozon", category_id, type_id
    if any(name.lower() in {"товары", "products"} for name in sheets):
        return "wb", None, None
    return None, None, None


def _int_or_none(value) -> int | None:
    try:
        return int(str(value).strip())
    except (TypeError, ValueError):
        return None
