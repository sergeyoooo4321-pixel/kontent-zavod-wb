from __future__ import annotations

import csv
import io
import json
import re
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from app.config import Settings
from app.models import ProductInput, ProductResult, SkuVariant
from app.template_fill import build_filled_templates, missing_templates_text


def build_sku_variants(product: ProductInput, settings: Settings) -> list[SkuVariant]:
    weight = product.weight_g or settings.DEFAULT_WEIGHT_G
    length = product.length_cm or settings.DEFAULT_LENGTH_CM
    width = product.width_cm or settings.DEFAULT_WIDTH_CM
    height = product.height_cm or settings.DEFAULT_HEIGHT_CM
    variants = []
    for qty in (1, 2, 3):
        sku = product.sku if qty == 1 else f"{product.sku}x{qty}"
        ozon_title = _ozon_title(product, qty)
        wb_short = _limit(_strip_brand(product.name, product.brand), 60)
        wb_full = _wb_full_title(product, qty)
        dims = _pack_dims(length, width, height, qty)
        variants.append(
            SkuVariant(
                sku=sku,
                qty=qty,
                ozon_title=ozon_title,
                wb_short_title=wb_short,
                wb_full_title=wb_full,
                weight_g=_round_weight(weight * qty),
                length_cm=dims[0],
                width_cm=dims[1],
                height_cm=dims[2],
            )
        )
    return variants


def build_zip(results: list[ProductResult], settings: Settings) -> bytes:
    links_csv = _links_csv(results)
    ozon = _workbook_bytes("ozon", results, settings)
    wb = _workbook_bytes("wb", results, settings)
    category_report = _category_report_bytes(results)
    marketplace_json = _marketplace_json(results)
    template_files, missing_templates = build_filled_templates(results, settings)
    readme = _readme(results)
    out = io.BytesIO()
    with ZipFile(out, "w", ZIP_DEFLATED) as zf:
        zf.writestr("links.csv", links_csv)
        zf.writestr("ozon.xlsx", ozon)
        zf.writestr("wildberries.xlsx", wb)
        zf.writestr("category_report.xlsx", category_report)
        zf.writestr("marketplace_fields.json", marketplace_json)
        zf.writestr("missing_templates.md", missing_templates_text(missing_templates))
        zf.writestr("README.txt", readme)
        for name, content in template_files.items():
            zf.writestr(name, content)
        for result in results:
            for image in result.images:
                if image.bytes_data and image.role != "source":
                    zf.writestr(f"photos/{result.input.sku}_{image.role}.jpg", image.bytes_data)
    return out.getvalue()


def _category_report_bytes(results: list[ProductResult]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Category report"
    ws.append(
        [
            "sku",
            "ozon_category_id",
            "ozon_type_id",
            "ozon_category_path",
            "wb_subject_id",
            "wb_subject_path",
            "missing_required",
            "warnings",
        ]
    )
    for result in results:
        profile = result.marketplace
        ozon = profile.ozon_category if profile else None
        wb_subject = profile.wb_subject if profile else None
        ws.append(
            [
                result.input.sku,
                ozon.id if ozon else "",
                ozon.type_id if ozon else "",
                ozon.path if ozon else "",
                wb_subject.id if wb_subject else "",
                wb_subject.path if wb_subject else "",
                "\n".join(profile.missing_required if profile else []),
                "\n".join((profile.warnings if profile else []) + result.warnings),
            ]
        )
    _style(ws)
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _marketplace_json(results: list[ProductResult]) -> str:
    return json.dumps(
        [
            {
                "sku": result.input.sku,
                "marketplace": result.marketplace.model_dump(mode="json") if result.marketplace else None,
            }
            for result in results
        ],
        ensure_ascii=False,
        indent=2,
    )


def _workbook_bytes(kind: str, results: list[ProductResult], settings: Settings) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Ozon" if kind == "ozon" else "Wildberries"
    headers = _ozon_headers() if kind == "ozon" else _wb_headers()
    ws.append(headers)
    for result in results:
        images = {img.role: img.url for img in result.images}
        profile = result.marketplace
        ozon_category = profile.ozon_category if profile else None
        wb_subject = profile.wb_subject if profile else None
        field_summary = _field_summary(profile.ozon_fields if kind == "ozon" and profile else profile.wb_fields if profile else [])
        missing_required = "\n".join(profile.missing_required if profile else [])
        for variant in build_sku_variants(result.input, settings):
            if kind == "ozon":
                ws.append(
                    [
                        variant.sku,
                        variant.ozon_title,
                        result.input.brand,
                        _description(result.input, variant.qty),
                        "0.22",
                        variant.weight_g,
                        variant.weight_g,
                        variant.length_cm * 10,
                        variant.width_cm * 10,
                        variant.height_cm * 10,
                        "mm",
                        result.input.price or settings.DEFAULT_PRICE,
                        images.get("main", ""),
                        images.get("pack2", ""),
                        images.get("pack3", ""),
                        images.get("extra", ""),
                        ozon_category.id if ozon_category else "",
                        ozon_category.type_id if ozon_category else "",
                        ozon_category.path if ozon_category else "",
                        field_summary,
                        missing_required,
                        result.input.extra,
                    ]
                )
            else:
                ws.append(
                    [
                        variant.sku,
                        variant.wb_short_title,
                        variant.wb_full_title,
                        result.input.brand,
                        _description(result.input, variant.qty),
                        variant.length_cm,
                        variant.width_cm,
                        variant.height_cm,
                        round(variant.weight_g / 1000, 3),
                        result.input.price or settings.DEFAULT_PRICE,
                        images.get("main", ""),
                        images.get("pack2", ""),
                        images.get("pack3", ""),
                        images.get("extra", ""),
                        wb_subject.id if wb_subject else "",
                        wb_subject.path if wb_subject else "",
                        field_summary,
                        missing_required,
                        result.input.extra,
                    ]
                )
    _style(ws)
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _ozon_headers() -> list[str]:
    return [
        "offer_id",
        "name",
        "brand",
        "description",
        "vat",
        "weight",
        "weight_packed_g",
        "depth_mm",
        "width_mm",
        "height_mm",
        "dimension_unit",
        "price",
        "primary_image",
        "image_pack2",
        "image_pack3",
        "image_extra",
        "description_category_id",
        "type_id",
        "category_path",
        "resolved_attributes",
        "missing_required",
        "comment",
    ]


def _wb_headers() -> list[str]:
    return [
        "vendorCode",
        "title_60",
        "full_title",
        "brand",
        "description",
        "length_cm",
        "width_cm",
        "height_cm",
        "weightBrutto_kg",
        "price",
        "media_main",
        "media_pack2",
        "media_pack3",
        "media_extra",
        "subject_id",
        "subject_path",
        "resolved_characteristics",
        "missing_required",
        "comment",
    ]


def _field_summary(fields) -> str:
    lines = []
    for field in fields:
        if field.value in (None, "", []):
            continue
        value = ", ".join(str(item) for item in field.value) if isinstance(field.value, list) else str(field.value)
        lines.append(f"{field.name}: {value}")
    return "\n".join(lines)


def _style(ws) -> None:
    fill = PatternFill("solid", fgColor="1F3A4A")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
    for column in ws.columns:
        letter = get_column_letter(column[0].column)
        width = min(max(len(str(cell.value or "")) for cell in column) + 2, 48)
        ws.column_dimensions[letter].width = width
    ws.freeze_panes = "A2"


def _links_csv(results: list[ProductResult]) -> str:
    out = io.StringIO()
    writer = csv.writer(out)
    writer.writerow(["sku", "photo_index", "role", "url"])
    for result in results:
        for image in result.images:
            writer.writerow([result.input.sku, result.input.photo_index, image.role, image.url])
    return out.getvalue()


def _readme(results: list[ProductResult]) -> str:
    lines = [
        "Content Zavod export pack",
        "",
        "Files:",
        "- photos/: generated JPG files",
        "- links.csv: SKU-to-image public links",
        "- ozon.xlsx: category-aware Ozon workbook",
        "- wildberries.xlsx: category-aware Wildberries workbook",
        "- category_report.xlsx: resolved categories and missing required fields",
        "- marketplace_fields.json: raw marketplace fields, allowed values and warnings",
        "- templates/: filled official cached XLSX templates when available",
        "- missing_templates.md: categories whose official XLSX templates are not cached",
        "",
        "Warnings:",
    ]
    warnings = [warning for result in results for warning in result.warnings]
    warnings.extend(warning for result in results if result.marketplace for warning in result.marketplace.warnings)
    warnings.extend(missing for result in results if result.marketplace for missing in result.marketplace.missing_required)
    lines.extend(f"- {warning}" for warning in warnings)
    if not warnings:
        lines.append("- none")
    return "\n".join(lines) + "\n"


def _description(product: ProductInput, qty: int) -> str:
    prefix = f"Набор {qty} шт. " if qty > 1 else ""
    brand = f"{product.brand} " if product.brand else ""
    facts = product.extra or "Подходит для ежедневного использования. Удобный формат для маркетплейса."
    return (
        f"{prefix}{brand}{product.name}. {facts} "
        "Товар подготовлен для карточки маркетплейса. Описание адаптировано под покупателя. "
        "Параметры и визуальные материалы проверяются оператором перед загрузкой."
    )


def _ozon_title(product: ProductInput, qty: int) -> str:
    base = " ".join(x for x in [product.brand, _strip_brand(product.name, product.brand)] if x).replace(" - ", " ")
    return base if qty == 1 else f"Набор {qty} шт {base}"


def _wb_full_title(product: ProductInput, qty: int) -> str:
    base = " ".join(x for x in [product.brand, _strip_brand(product.name, product.brand)] if x)
    return base if qty == 1 else f"Набор {qty} шт {base}"


def _strip_brand(name: str, brand: str) -> str:
    value = name.strip()
    if brand and value.lower().startswith(brand.lower()):
        value = value[len(brand) :].strip(" -")
    return value


def _limit(value: str, limit: int) -> str:
    value = re.sub(r"\s+", " ", value).strip()
    if len(value) <= limit:
        return value
    cut = value[: limit - 1].rsplit(" ", 1)[0]
    return f"{cut}…" if cut else value[: limit - 1] + "…"


def _pack_dims(length: int, width: int, height: int, qty: int) -> tuple[int, int, int]:
    dims = [length, width, height]
    if qty > 1:
        idx = min(range(3), key=lambda i: dims[i])
        dims[idx] *= qty
    return tuple(int(x) for x in dims)  # type: ignore[return-value]


def _round_weight(value: int) -> int:
    if value <= 100:
        return int(((value + 9) // 10) * 10)
    return int(((value + 49) // 50) * 50)
