from __future__ import annotations

import io
import re
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.config import Settings
from app.models import MarketplaceFieldValue, ProductResult, SkuVariant
from app.template_cache import find_template


def build_filled_templates(results: list[ProductResult], settings: Settings) -> tuple[dict[str, bytes], list[str]]:
    files: dict[str, bytes] = {}
    missing: list[str] = []
    grouped = _group_by_marketplace_category(results)
    for key, group in grouped.items():
        marketplace, category_id, type_id = key
        cached = find_template(settings, marketplace, category_id, type_id)
        if cached is None:
            category_label = _category_label(group[0], marketplace)
            missing.append(f"{marketplace.upper()} category={category_id} type={type_id or ''} {category_label}".strip())
            continue
        files[f"templates/{marketplace}_{category_id}{'_' + str(type_id) if type_id else ''}.xlsx"] = _fill_template(cached.path, group, settings, marketplace)
    return files, missing


def missing_templates_text(missing: list[str]) -> str:
    if not missing:
        return "All category templates were found in cache.\n"
    lines = [
        "Missing official XLSX templates",
        "",
        "The bot resolved marketplace categories, but an official cabinet template is not cached for these categories.",
        "Send the empty XLSX template to the bot with a caption like:",
        "template ozon <description_category_id> <type_id>",
        "template wb <subject_id>",
        "",
        "Missing:",
    ]
    lines.extend(f"- {item}" for item in missing)
    return "\n".join(lines) + "\n"


def _group_by_marketplace_category(results: list[ProductResult]) -> dict[tuple[str, int, int | None], list[ProductResult]]:
    grouped: dict[tuple[str, int, int | None], list[ProductResult]] = defaultdict(list)
    for result in results:
        profile = result.marketplace
        if not profile:
            continue
        if profile.ozon_category:
            grouped[("ozon", profile.ozon_category.id, profile.ozon_category.type_id)].append(result)
        if profile.wb_subject:
            grouped[("wb", profile.wb_subject.id, None)].append(result)
    return grouped


def _fill_template(path: Path, results: list[ProductResult], settings: Settings, marketplace: str) -> bytes:
    from app.excel_export import build_sku_variants

    workbook = load_workbook(path)
    sheet_name, header_row, data_start_row = _detect_sheet(workbook.sheetnames, marketplace)
    ws = workbook[sheet_name]
    headers = {col: str(ws.cell(header_row, col).value or "").strip() for col in range(1, ws.max_column + 1)}
    row_idx = data_start_row
    for result in results:
        for variant in build_sku_variants(result.input, settings):
            row_values = _values_for_variant(result, variant, marketplace)
            for col, header in headers.items():
                value = _resolve_header_value(header, row_values)
                if value is not None:
                    ws.cell(row_idx, col).value = value
            row_idx += 1
    out = io.BytesIO()
    workbook.save(out)
    return out.getvalue()


def _detect_sheet(sheetnames: list[str], marketplace: str) -> tuple[str, int, int]:
    if marketplace == "ozon":
        if "Шаблон" in sheetnames:
            return "Шаблон", 2, 5
        if "Template" in sheetnames:
            return "Template", 2, 5
    if marketplace == "wb":
        if "Товары" in sheetnames:
            return "Товары", 3, 5
        if "Products" in sheetnames:
            return "Products", 3, 5
    return sheetnames[0], 1, 2


def _values_for_variant(result: ProductResult, variant: SkuVariant, marketplace: str) -> dict[str, Any]:
    product = result.input
    images = {image.role: image.url for image in result.images}
    profile = result.marketplace
    fields: list[MarketplaceFieldValue] = []
    category_path = ""
    if profile:
        if marketplace == "ozon":
            fields = profile.ozon_fields
            category_path = profile.ozon_category.path if profile.ozon_category else ""
        else:
            fields = profile.wb_fields
            category_path = profile.wb_subject.path if profile.wb_subject else ""
    dynamic = {_norm(field.name): _format_value(field.value) for field in fields if field.value not in (None, "", [])}
    title = variant.ozon_title if marketplace == "ozon" else variant.wb_full_title
    return {
        "sku": variant.sku,
        "title": title,
        "name": title,
        "brand": product.brand,
        "description": _description(product, variant.qty),
        "category": category_path,
        "weight_g": variant.weight_g,
        "weight_kg": round(variant.weight_g / 1000, 3),
        "length_cm": variant.length_cm,
        "width_cm": variant.width_cm,
        "height_cm": variant.height_cm,
        "length_mm": variant.length_cm * 10,
        "width_mm": variant.width_cm * 10,
        "height_mm": variant.height_cm * 10,
        "price": product.price or "",
        "vat": "22" if marketplace == "wb" else "0.22",
        "image": images.get("main", ""),
        "images": ", ".join(url for role, url in images.items() if role != "source"),
        **dynamic,
    }


def _resolve_header_value(header: str, values: dict[str, Any]) -> Any:
    norm = _norm(header)
    if not norm:
        return None
    direct = values.get(norm)
    if direct is not None:
        return direct
    rules = [
        (("offer", "артикул", "vendorcode", "vendor code"), "sku"),
        (("название", "наименование", "name", "title"), "title"),
        (("бренд", "brand"), "brand"),
        (("описание", "аннотация", "description"), "description"),
        (("категория", "category"), "category"),
        (("ндс", "vat"), "vat"),
        (("цена", "price"), "price"),
        (("фото", "изображ", "картин", "image", "photo"), "image"),
        (("вес брутто", "вес в упаков", "weight packed"), "weight_g"),
        (("вес", "weight"), "weight_g"),
        (("длина", "length", "глубина", "depth"), "length_mm" if "мм" in norm or "mm" in norm else "length_cm"),
        (("ширина", "width"), "width_mm" if "мм" in norm or "mm" in norm else "width_cm"),
        (("высота", "height"), "height_mm" if "мм" in norm or "mm" in norm else "height_cm"),
        (("штрих", "barcode"), "barcode"),
    ]
    for needles, key in rules:
        if any(needle in norm for needle in needles):
            return "" if key == "barcode" else values.get(key)
    return None


def _category_label(result: ProductResult, marketplace: str) -> str:
    profile = result.marketplace
    if not profile:
        return ""
    category = profile.ozon_category if marketplace == "ozon" else profile.wb_subject
    return category.path if category else ""


def _description(product, qty: int) -> str:
    prefix = f"Набор {qty} шт. " if qty > 1 else ""
    return f"{prefix}{product.brand + ' ' if product.brand else ''}{product.name}. {product.extra or ''}".strip()


def _format_value(value: Any) -> Any:
    if isinstance(value, list):
        return ", ".join(str(item) for item in value)
    return value


def _norm(value: str) -> str:
    return re.sub(r"\s+", " ", re.sub(r"[*:]+", "", value.lower())).strip()
