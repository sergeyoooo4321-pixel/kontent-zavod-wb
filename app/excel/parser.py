"""Парсер xlsx-шаблонов Ozon и Wildberries в единую структуру TemplateSpec.

Поддерживает 2 формата:
    - **Ozon**: листы ['Шаблон', 'configs', 'info', 'validation', ...]
              header_row=2, required_row=3, hint_row=4, data_start=5
              dropdown'ы хранятся как named ranges (`name5`, `name6`, ...)
              на листе `validation`

    - **Wildberries**: листы ['Товары', 'Инструкция']
              header_row=3, hint_row=4, data_start=5
              ширина листа ~3600 столбцов (одна категория = много характеристик)
              dropdown'ы — DataValidation с inline-списками либо на лист Товары

Ozon-шаблоны иногда падают на normal load из-за DataValidation в формате
MultiCellRange (баг openpyxl) — fallback на read_only=True (теряем DataValidation,
но получаем структуру полей).
"""
from __future__ import annotations

import logging
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterable

import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)


# ─── Public types ────────────────────────────────────────────────


@dataclass
class TemplateField:
    """Одно поле шаблона."""
    name: str
    column: int                 # 1-based номер столбца на data-листе
    required: bool = False
    description: str = ""
    dropdown: list[str] | None = None  # None если поле свободного ввода
    # WB-специфика: charcID и charcType из defined_names (characteristic_<id>_<type>)
    # charcType: 0=number, 1=string, 4=dictionary_single, 5=dictionary_multi
    wb_charc_id: int | None = None
    wb_charc_type: int | None = None
    wb_static: bool = False     # true для static_* полей (служебные WB-колонки)


@dataclass
class TemplateSpec:
    """Структура распаршенного шаблона."""
    marketplace: str            # "ozon" | "wb"
    sheet_name: str             # имя листа с данными
    header_row: int             # 1-based номер строки с именами полей
    data_start_row: int         # 1-based с какой строки начинать данные
    fields: list[TemplateField] = field(default_factory=list)
    category_id: int | None = None       # Ozon: DESCRIPTION_CATEGORY_ID из configs
    type_id: int | None = None           # Ozon: TYPE_ID если найден
    raw_path: str = ""                   # путь исходного файла
    parse_warnings: list[str] = field(default_factory=list)


# ─── Format detection ────────────────────────────────────────────


def detect_format(wb: Workbook) -> str:
    """Возвращает 'ozon' | 'wb' | 'unknown' по составу листов."""
    sheets = set(wb.sheetnames)
    if "Шаблон" in sheets and any(s.lower().startswith("validation") or s == "validation"
                                  for s in sheets):
        return "ozon"
    if "Товары" in sheets and "Инструкция" in sheets:
        return "wb"
    if "Шаблон" in sheets:
        return "ozon"
    if "Товары" in sheets:
        return "wb"
    return "unknown"


# ─── Loader with fallback ────────────────────────────────────────


def _load_wb(path: Path | str) -> tuple[Workbook, bool]:
    """Открывает workbook. Returns (wb, used_readonly).

    Сначала пытается полную загрузку. Если openpyxl падает на DataValidation
    (известный баг с MultiCellRange) — fallback на read_only=True (теряем
    Data Validation, но получаем структуру).
    """
    try:
        return openpyxl.load_workbook(path, data_only=False), False
    except Exception as e:
        logger.warning("normal load failed (%s), retry read_only=True", e)
        return openpyxl.load_workbook(path, data_only=False, read_only=True), True


# ─── Ozon parser ─────────────────────────────────────────────────


_NAMED_RANGE_RE = re.compile(r"^[a-zA-Z_][\w]*$")


def _parse_ozon(wb: Workbook, used_readonly: bool, warnings: list[str]) -> TemplateSpec:
    sheet_name = "Шаблон" if "Шаблон" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]
    header_row = 2
    required_row = 3
    hint_row = 4
    data_start_row = 5

    # Конкретные значения из named ranges на листе validation
    dropdowns_by_col = _ozon_dropdowns_by_column(wb, ws, used_readonly, warnings)

    fields: list[TemplateField] = []
    max_col = ws.max_column or 1
    for col in range(1, max_col + 1):
        name = _cell_str(ws, header_row, col)
        if not name:
            continue
        # маркер required: либо `*` в имени, либо «Обязательное поле» в row 3
        required_marker = _cell_str(ws, required_row, col).lower()
        required = name.endswith("*") or "обязател" in required_marker
        # очистка имени от `*`
        clean_name = name.rstrip("*").strip()
        hint = _cell_str(ws, hint_row, col)
        fields.append(TemplateField(
            name=clean_name,
            column=col,
            required=required,
            description=hint,
            dropdown=dropdowns_by_col.get(col),
        ))

    cat_id, type_id = _ozon_category_ids_from_configs(wb, warnings)

    return TemplateSpec(
        marketplace="ozon",
        sheet_name=sheet_name,
        header_row=header_row,
        data_start_row=data_start_row,
        fields=fields,
        category_id=cat_id,
        type_id=type_id,
        parse_warnings=warnings,
    )


def _ozon_dropdowns_by_column(
    wb: Workbook,
    ws: Worksheet,
    used_readonly: bool,
    warnings: list[str],
) -> dict[int, list[str]]:
    """Извлекает list-DataValidation в формате `=name5` и резолвит named ranges
    на листе validation."""
    out: dict[int, list[str]] = {}
    if used_readonly:
        # read_only mode не даёт data_validations
        warnings.append("ozon: dropdown'ы недоступны в read_only-режиме")
        return out

    try:
        dv_list = list(ws.data_validations.dataValidation)
    except Exception as e:
        warnings.append(f"ozon: data_validations parse failed: {e}")
        return out

    # Резолвер named ranges → значения с листа validation
    val_ws: Worksheet | None = None
    if "validation" in wb.sheetnames:
        val_ws = wb["validation"]

    name_to_col: dict[str, int] = {}
    if val_ws is not None:
        # На листе validation первая строка может быть заголовком, либо данные с row=1.
        # Named ranges типа `name5`, `name6` указывают на колонки.
        # Без точного резолва (defined_names) пытаемся прочитать named ranges из workbook.
        try:
            for dn_name in wb.defined_names:
                dn = wb.defined_names[dn_name]
                if not dn or not dn.value:
                    continue
                # value формата "validation!$F$2:$F$200"
                val = str(dn.value)
                if "!" not in val:
                    continue
                _, rng = val.split("!", 1)
                m = re.match(r"\$?([A-Z]+)\$?\d+:\$?([A-Z]+)\$?\d+", rng) or \
                    re.match(r"\$?([A-Z]+)\$?\d+", rng)
                if m:
                    col_letter = m.group(1)
                    name_to_col[dn_name] = _col_letter_to_index(col_letter)
        except Exception as e:
            warnings.append(f"ozon: defined_names parse failed: {e}")

    for dv in dv_list:
        if dv.type != "list" or not dv.formula1:
            continue
        f = (dv.formula1 or "").strip().lstrip("=")
        values: list[str] = []
        if f.startswith('"') and f.endswith('"'):
            values = [s.strip() for s in f[1:-1].split(",") if s.strip()]
        elif _NAMED_RANGE_RE.match(f) and val_ws is not None and f in name_to_col:
            col_idx = name_to_col[f]
            for row in val_ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=True):
                v = row[0]
                if v is None:
                    continue
                s = str(v).strip()
                if s and s not in values:
                    values.append(s)
        elif "!" in f:
            sheet_part, rng = f.split("!", 1)
            sn = sheet_part.strip("'\"")
            if sn in wb.sheetnames:
                try:
                    src = wb[sn]
                    for row in src[rng]:
                        for c in row:
                            if c.value is None:
                                continue
                            s = str(c.value).strip()
                            if s and s not in values:
                                values.append(s)
                except Exception:
                    pass

        if not values:
            continue
        # к каким столбцам data-листа применяется
        for sqref in (str(dv.sqref) or "").split():
            for c in _expand_columns(sqref):
                # дедуп: если у столбца уже есть значения — мерджим
                bucket = out.setdefault(c, [])
                for v in values:
                    if v not in bucket:
                        bucket.append(v)
    return out


def _ozon_category_ids_from_configs(wb: Workbook, warnings: list[str]) -> tuple[int | None, int | None]:
    """Возвращает (category_id, type_id) из листа `configs`, если есть."""
    if "configs" not in wb.sheetnames:
        return None, None
    ws = wb["configs"]
    cat_id: int | None = None
    type_id: int | None = None
    try:
        for r in range(1, min(20, ws.max_row + 1)):
            key = _cell_str(ws, r, 1).upper()
            val = _cell_str(ws, r, 2)
            if not key or not val:
                continue
            if key == "DESCRIPTION_CATEGORY_ID":
                try:
                    cat_id = int(val)
                except ValueError:
                    pass
            elif key in ("TYPE_ID", "PRODUCTS_TYPE_ID"):
                try:
                    type_id = int(val)
                except ValueError:
                    pass
    except Exception as e:
        warnings.append(f"ozon: configs parse failed: {e}")
    return cat_id, type_id


# ─── WB parser ──────────────────────────────────────────────────


_WB_CHARC_NAME_RE = re.compile(r"^characteristic_(\d+)_(\d+)$")
_WB_STATIC_NAME_RE = re.compile(r"^static_(\w+)_(\d+)$")


def _wb_defined_names_by_column(wb: Workbook, sheet_name: str) -> dict[int, dict]:
    """Парсит defined_names типа `characteristic_<id>_<type>` и `static_*_<type>`,
    возвращает {column_index: {charc_id, charc_type, static_name?}}.

    WB-шаблоны хранят charcID и charcType (0/1/4/5) через named ranges,
    указывающие на ячейку Товары!$F$3 (имя поля). Это единственный способ
    узнать тип поля — DataValidation в WB-xlsx нет.
    """
    out: dict[int, dict] = {}
    try:
        for dn_name in wb.defined_names:
            try:
                dn = wb.defined_names[dn_name]
                val = (dn.value or "") if dn else ""
            except Exception:
                continue
            if not val or "!" not in val:
                continue
            sn, rng = val.split("!", 1)
            sn = sn.strip("'\"")
            if sn != sheet_name:
                continue
            m = re.match(r"\$?([A-Z]+)\$?(\d+)", rng)
            if not m:
                continue
            col = _col_letter_to_index(m.group(1))

            mc = _WB_CHARC_NAME_RE.match(dn_name)
            if mc:
                out[col] = {
                    "charc_id": int(mc.group(1)),
                    "charc_type": int(mc.group(2)),
                    "static": False,
                }
                continue
            ms = _WB_STATIC_NAME_RE.match(dn_name)
            if ms:
                out[col] = {
                    "charc_id": None,
                    "charc_type": int(ms.group(2)),
                    "static": True,
                    "static_name": ms.group(1),
                }
    except Exception as e:
        logger.warning("wb defined_names parse failed: %s", e)
    return out


def _parse_wb(wb: Workbook, used_readonly: bool, warnings: list[str]) -> TemplateSpec:
    sheet_name = "Товары" if "Товары" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]
    header_row = 3
    hint_row = 4
    data_start_row = 5

    # обязательные поля у WB определяются по описанию и по жёсткому списку
    REQUIRED_HINT_KEYWORDS = ("обязател",)
    HARD_REQUIRED_NAMES = {
        "артикул продавца", "наименование", "категория продавца", "бренд",
        "описание",
    }

    dropdowns_by_col = _wb_dropdowns_by_column(ws, used_readonly, warnings)
    charcs_by_col = _wb_defined_names_by_column(wb, sheet_name) if not used_readonly else {}
    if not charcs_by_col and not used_readonly:
        warnings.append("wb: defined_names с charcID не найдены")

    fields: list[TemplateField] = []
    max_col = ws.max_column or 1
    seen_cols = 0
    empty_streak = 0
    for col in range(1, max_col + 1):
        name = _cell_str(ws, header_row, col)
        if not name:
            empty_streak += 1
            if empty_streak > 200:
                break
            continue
        empty_streak = 0
        seen_cols += 1
        hint = _cell_str(ws, hint_row, col)
        required = (
            name.lower().rstrip("*").strip() in HARD_REQUIRED_NAMES
            or any(kw in hint.lower() for kw in REQUIRED_HINT_KEYWORDS)
            or name.endswith("*")
        )
        clean_name = name.rstrip("*").strip()
        meta = charcs_by_col.get(col, {})
        fields.append(TemplateField(
            name=clean_name,
            column=col,
            required=required,
            description=hint,
            dropdown=dropdowns_by_col.get(col),
            wb_charc_id=meta.get("charc_id"),
            wb_charc_type=meta.get("charc_type"),
            wb_static=bool(meta.get("static")),
        ))

    return TemplateSpec(
        marketplace="wb",
        sheet_name=sheet_name,
        header_row=header_row,
        data_start_row=data_start_row,
        fields=fields,
        parse_warnings=warnings,
    )


def _wb_dropdowns_by_column(
    ws: Worksheet,
    used_readonly: bool,
    warnings: list[str],
) -> dict[int, list[str]]:
    """WB обычно использует inline-списки в формуле DataValidation."""
    out: dict[int, list[str]] = {}
    if used_readonly:
        warnings.append("wb: dropdown'ы недоступны в read_only-режиме")
        return out
    try:
        dv_list = list(ws.data_validations.dataValidation)
    except Exception as e:
        warnings.append(f"wb: data_validations parse failed: {e}")
        return out
    for dv in dv_list:
        if dv.type != "list" or not dv.formula1:
            continue
        f = (dv.formula1 or "").strip().lstrip("=")
        values: list[str] = []
        if f.startswith('"') and f.endswith('"'):
            values = [s.strip() for s in f[1:-1].split(",") if s.strip()]
        if not values:
            continue
        for sqref in (str(dv.sqref) or "").split():
            for c in _expand_columns(sqref):
                bucket = out.setdefault(c, [])
                for v in values:
                    if v not in bucket:
                        bucket.append(v)
    return out


# ─── Public entry ────────────────────────────────────────────────


def parse_template(path: str | Path) -> TemplateSpec:
    """Главный entry point. Авто-детект формата + парсинг."""
    p = Path(path)
    wb, used_readonly = _load_wb(p)
    warnings: list[str] = []
    if used_readonly:
        warnings.append("opened in read_only=True (DataValidation skipped)")

    fmt = detect_format(wb)
    if fmt == "ozon":
        spec = _parse_ozon(wb, used_readonly, warnings)
    elif fmt == "wb":
        spec = _parse_wb(wb, used_readonly, warnings)
    else:
        raise ValueError(f"unknown template format, sheets={wb.sheetnames!r}")
    spec.raw_path = str(p)
    return spec


# ─── helpers ────────────────────────────────────────────────────


def _cell_str(ws: Worksheet, row: int, col: int) -> str:
    if row < 1 or col < 1:
        return ""
    try:
        v = ws.cell(row, col).value
    except Exception:
        return ""
    return str(v).strip() if v is not None else ""


_COL_CELL_RE = re.compile(r"\$?([A-Z]+)\$?(\d+)")


def _col_letter_to_index(letter: str) -> int:
    n = 0
    for ch in letter.upper():
        if "A" <= ch <= "Z":
            n = n * 26 + (ord(ch) - ord("A") + 1)
    return n


def _expand_columns(cell_range: str) -> Iterable[int]:
    """Из 'A5:A500002' возвращает [1]; из 'B5:D5' возвращает [2,3,4]."""
    if ":" not in cell_range:
        m = _COL_CELL_RE.match(cell_range)
        if m:
            yield _col_letter_to_index(m.group(1))
        return
    a, b = cell_range.split(":", 1)
    ma, mb = _COL_CELL_RE.match(a), _COL_CELL_RE.match(b)
    if not ma or not mb:
        return
    c1 = _col_letter_to_index(ma.group(1))
    c2 = _col_letter_to_index(mb.group(1))
    for c in range(c1, c2 + 1):
        yield c
