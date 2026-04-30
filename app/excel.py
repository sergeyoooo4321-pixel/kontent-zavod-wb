"""Парсинг XLSX-шаблонов Ozon: считывание справочников через Data Validation,
заполнение строк, сохранение."""
from __future__ import annotations

import io
import logging
import re
from typing import Any

from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


_CELL_RANGE_RE = re.compile(r"([A-Z]+)(\d+)")


def _expand_range(cell_range: str) -> list[tuple[int, int]]:
    """'A1:A5' → [(1,1),(2,1),(3,1),(4,1),(5,1)]."""
    if ":" not in cell_range:
        m = _CELL_RANGE_RE.match(cell_range)
        if not m:
            return []
        col = _col_letter_to_index(m.group(1))
        return [(int(m.group(2)), col)]
    a, b = cell_range.split(":", 1)
    ma, mb = _CELL_RANGE_RE.match(a), _CELL_RANGE_RE.match(b)
    if not ma or not mb:
        return []
    c1, r1 = _col_letter_to_index(ma.group(1)), int(ma.group(2))
    c2, r2 = _col_letter_to_index(mb.group(1)), int(mb.group(2))
    cells = []
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            cells.append((r, c))
    return cells


def _col_letter_to_index(letter: str) -> int:
    n = 0
    for ch in letter:
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n


class OzonTemplate:
    """Обёртка над workbook XLSX-шаблона Ozon."""

    def __init__(self, wb: Workbook, sheet_name: str | None = None):
        self._wb = wb
        # Лист с шаблоном — обычно первый или с именем 'Шаблон'/'Template'
        if sheet_name and sheet_name in wb.sheetnames:
            self._sheet = wb[sheet_name]
        elif "Шаблон" in wb.sheetnames:
            self._sheet = wb["Шаблон"]
        elif "Template" in wb.sheetnames:
            self._sheet = wb["Template"]
        else:
            self._sheet = wb.active
        self._headers: list[str] | None = None

    @classmethod
    def from_bytes(cls, data: bytes, sheet_name: str | None = None) -> "OzonTemplate":
        wb = load_workbook(io.BytesIO(data), data_only=False)
        return cls(wb, sheet_name)

    @classmethod
    def from_file(cls, path: str, sheet_name: str | None = None) -> "OzonTemplate":
        wb = load_workbook(path, data_only=False)
        return cls(wb, sheet_name)

    # ─── headers ──────────────────────────────────────────

    def headers(self, header_row: int = 1) -> list[str]:
        """Заголовки колонок (по умолчанию 1-я строка)."""
        if self._headers is None:
            row = self._sheet[header_row]
            self._headers = [str(c.value or "").strip() for c in row]
        return self._headers

    # ─── справочники ──────────────────────────────────────

    def read_dictionaries(self, header_row: int = 1) -> dict[str, list[str]]:
        """Возвращает {column_header: [allowed_values]} для колонок с Data Validation типа list.

        Также пытается загрузить отдельный лист 'Справочники'/'Dictionaries' если есть.
        """
        result: dict[str, list[str]] = {}
        headers = self.headers(header_row)

        # 1) Data Validation на основном листе
        for dv in self._sheet.data_validations.dataValidation:
            if dv.type != "list" or not dv.formula1:
                continue
            values = self._extract_list_values(dv.formula1)
            if not values:
                continue
            # к каким колонкам применяется
            for sqref in (dv.sqref or "").split():
                cells = _expand_range(sqref)
                cols = {c for _, c in cells}
                for col_idx in cols:
                    if col_idx <= len(headers):
                        header = headers[col_idx - 1]
                        if header:
                            existing = result.setdefault(header, [])
                            for v in values:
                                if v not in existing:
                                    existing.append(v)

        # 2) Лист со справочниками — мерджим если есть
        for sheet_name in ("Справочники", "Dictionaries", "Lists"):
            if sheet_name in self._wb.sheetnames:
                ws = self._wb[sheet_name]
                first_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))
                if not first_row:
                    continue
                col_headers = [str(v or "").strip() for v in first_row[0]]
                for col_idx, header in enumerate(col_headers, 1):
                    if not header:
                        continue
                    values: list[str] = []
                    for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx, values_only=True):
                        v = row[0]
                        if v is None:
                            continue
                        s = str(v).strip()
                        if s and s not in values:
                            values.append(s)
                    if values:
                        existing = result.setdefault(header, [])
                        for v in values:
                            if v not in existing:
                                existing.append(v)

        return result

    def _extract_list_values(self, formula: str) -> list[str]:
        """Извлекает значения из формулы Data Validation.

        V5 — теперь поддерживает:
        - Inline: `"a,b,c"` → ['a','b','c']
        - Reference: `=Sheet!$A$1:$A$100` — читает диапазон с указанного листа
        """
        f = (formula or "").strip().lstrip("=")
        if f.startswith('"') and f.endswith('"'):
            inner = f[1:-1]
            return [s.strip() for s in inner.split(",") if s.strip()]
        if "!" in f:
            sheet_part, rng = f.split("!", 1)
            sheet_name = sheet_part.strip("'\"")
            if sheet_name in self._wb.sheetnames:
                try:
                    ws = self._wb[sheet_name]
                    out: list[str] = []
                    for row in ws[rng]:
                        for c in row:
                            if c.value is None:
                                continue
                            s = str(c.value).strip()
                            if s and s not in out:
                                out.append(s)
                    return out
                except Exception as e:
                    logger.warning("DV reference parse %s: %s", f, e)
                    return []
            else:
                logger.warning("DV references missing sheet %s", sheet_name)
        return []

    # ─── заполнение ───────────────────────────────────────

    def fill_rows(
        self,
        rows: list[dict[str, Any]],
        *,
        header_row: int = 1,
        start_row: int = 2,
    ) -> None:
        """Заполняет строки начиная со start_row. rows — список словарей по headers."""
        headers = self.headers(header_row)
        for r_idx, row in enumerate(rows):
            for col_idx, header in enumerate(headers, 1):
                if header in row:
                    self._sheet.cell(row=start_row + r_idx, column=col_idx, value=row[header])

    def to_bytes(self) -> bytes:
        buf = io.BytesIO()
        self._wb.save(buf)
        return buf.getvalue()

    def save(self, path: str) -> None:
        self._wb.save(path)
