"""Internal API для гнома (cz-gnome.service на :8001).

Эндпоинты под защитой X-Internal-Token. Гном дёргает их HTTP'ом, чтобы
переиспользовать уже поднятые в cz-backend клиенты (kie, s3, wb, ozon)
без второго экземпляра конфигов и сетевых сессий.
"""
from __future__ import annotations

import logging
import time
import uuid
from typing import Any

from fastapi import APIRouter, Header, HTTPException, Request
from pydantic import BaseModel

from .config import settings
from .pipeline import Deps
from .prompts import (
    build_design_director_system,
    build_design_director_user,
    compile_image_prompt,
    build_extra_prompt,
    build_main_prompt,
    build_pack_prompt,
)
from .s3 import S3Client

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/internal", tags=["internal"])


def _check_token(token: str | None) -> None:
    if settings.INTERNAL_TOKEN and token != settings.INTERNAL_TOKEN:
        raise HTTPException(status_code=403, detail="bad internal token")


def _pick_wb_client(cabinet_name: str | None, deps: "Deps"):
    """Выбрать WBClient по fallback-цепочке: указанный кабинет → default → любой
    кабинет с WB_TOKEN. Справочники WB — глобальный каталог, любой валидный
    токен подходит, поэтому fallback безопасен.

    Возвращает (WBClient | None, used_label).
    """
    from .pipeline import _wb_client_for

    # 1. Конкретный кабинет
    if cabinet_name:
        cab = settings.get_cabinet(cabinet_name)
        if cab is not None and cab.has_wb:
            client = _wb_client_for(cab, deps)
            if client is not None:
                return client, cab.label

    # 2. Default
    if deps.wb is not None:
        return deps.wb, "default"

    # 3. Любой кабинет с WB
    for cab in settings.list_cabinets():
        if cab.has_wb:
            client = _wb_client_for(cab, deps)
            if client is not None:
                return client, cab.label

    return None, ""


def _pick_ozon_client(cabinet_name: str | None, deps: "Deps"):
    """То же что _pick_wb_client, но для Ozon. Категории Ozon — глобальный каталог."""
    from .pipeline import _ozon_client_for

    if cabinet_name:
        cab = settings.get_cabinet(cabinet_name)
        if cab is not None and cab.has_ozon:
            client = _ozon_client_for(cab, deps)
            if client is not None:
                return client, cab.label

    if deps.ozon is not None:
        return deps.ozon, "default"

    for cab in settings.list_cabinets():
        if cab.has_ozon:
            client = _ozon_client_for(cab, deps)
            if client is not None:
                return client, cab.label

    return None, ""


# ─── /internal/generate_image ──────────────────────────────────────


class GenImageIn(BaseModel):
    src_url: str  # публичный URL исходной фотки (юзер прислал, уже в S3)
    brand: str
    name: str
    sku: str = ""  # опционально для tagging в S3
    variants: list[str] = ["main", "pack2", "pack3", "extra"]


class GenImageOut(BaseModel):
    ok: bool
    images: dict[str, str]  # tag → public_url
    errors: dict[str, str]  # tag → error msg
    brief: dict | None = None


@router.post("/generate_image", response_model=GenImageOut)
async def gen_image(
    req: GenImageIn,
    request: Request,
    x_internal_token: str | None = Header(default=None, alias="X-Internal-Token"),
):
    _check_token(x_internal_token)
    deps: Deps = request.app.state.deps
    sku = req.sku or uuid.uuid4().hex[:8]
    batch_id = f"gnome-{int(time.time())}-{uuid.uuid4().hex[:6]}"

    # 1. brief через vision LLM (как в pipeline.process_product_images)
    brief: dict = {}
    try:
        brief = await deps.kie.chat_json_with_vision(
            system=build_design_director_system(),
            user=build_design_director_user(req.name, req.brand),
            image_url=req.src_url,
        )
        logger.info("internal/gen_image brief sku=%s ok", sku)
    except Exception as e:
        logger.warning("internal/gen_image brief failed sku=%s: %s — generic", sku, e)
        brief = {}

    # 2. Параллельно все варианты
    import asyncio
    images: dict[str, str] = {}
    errors: dict[str, str] = {}

    async def _gen_one(tag: str) -> tuple[str, str | None, str | None]:
        try:
            # compile_image_prompt поддерживает только main / pack2 / pack3 / extra
            qty_for_pack = 2 if tag == "pack2" else (3 if tag == "pack3" else None)
            if tag in ("main", "extra", "pack2", "pack3"):
                prompt_kwargs = {"mode": tag}
                if qty_for_pack:
                    prompt_kwargs["qty"] = qty_for_pack
                prompt = compile_image_prompt(brief, req.name, **prompt_kwargs)
            else:
                # незнакомый tag — отдадим main-промпт чтобы не упасть
                prompt = compile_image_prompt(brief, req.name, mode="main")
            kie_url = await deps.kie.generate_image_with_retry(
                prompt=prompt,
                input_urls=[req.src_url],
            )
            # aitunnel часто возвращает data:URI / base64 вместо URL —
            # fetch_or_decode_image универсально обработает все варианты.
            content = await deps.kie.fetch_or_decode_image(kie_url)
            public = await deps.s3.put_public(
                S3Client.build_key(batch_id, sku, tag), content,
            )
            return tag, public, None
        except Exception as e:
            return tag, None, str(e)[:300]

    results = await asyncio.gather(*[_gen_one(t) for t in req.variants])
    for tag, url, err in results:
        if url:
            images[tag] = url
        else:
            errors[tag] = err or "unknown"

    return GenImageOut(
        ok=bool(images),
        images=images,
        errors=errors,
        brief=brief or None,
    )


# ─── /internal/match_category ──────────────────────────────────────


class MatchCategoryIn(BaseModel):
    name: str
    brand: str = ""
    main_image_url: str | None = None
    side: str = "both"  # "ozon" | "wb" | "both"


class MatchCategoryOut(BaseModel):
    ok: bool
    ozon: list[dict] | None = None
    wb: list[dict] | None = None
    error: str | None = None


@router.post("/match_category", response_model=MatchCategoryOut)
async def match_category(
    req: MatchCategoryIn,
    request: Request,
    x_internal_token: str | None = Header(default=None, alias="X-Internal-Token"),
):
    """Stub: полный match_category — внутри make_batch_zip; этот endpoint
    deprecated.
    """
    _check_token(x_internal_token)
    return MatchCategoryOut(
        ok=False,
        ozon=None,
        wb=None,
        error=(
            "match_category endpoint deprecated — категории матчатся внутри "
            "make_batch_zip автоматически. Не вызывай отдельно."
        ),
    )


# ─── /internal/fill_card ───────────────────────────────────────────


class FillCardIn(BaseModel):
    sku: str
    brand: str
    name: str
    images: dict[str, str]  # tag → public_url
    cabinet: str | None = None  # имя кабинета или None = default
    dry_run: bool = True


class FillCardOut(BaseModel):
    ok: bool
    dry_run: bool
    payload: dict | None = None
    error: str | None = None


@router.post("/fill_card", response_model=FillCardOut)
async def fill_card(
    req: FillCardIn,
    request: Request,
    x_internal_token: str | None = Header(default=None, alias="X-Internal-Token"),
):
    """MVP: собирает базовый payload карточки. Полная заливка с атрибутами —
    через старый кнопочный pipeline (где есть категории, справочники, валидация).
    """
    _check_token(x_internal_token)
    payload = {
        "sku": req.sku,
        "brand": req.brand,
        "name": req.name,
        "images": req.images,
        "cabinet": req.cabinet or settings.default_cabinet_name,
        "note": "MVP-payload от гнома. Полная заливка — через «📦 Новая партия».",
    }
    if req.dry_run:
        return FillCardOut(ok=True, dry_run=True, payload=payload)
    return FillCardOut(
        ok=False,
        dry_run=False,
        error="Полная заливка через гнома пока не реализована. "
              "Используй DRY_RUN=true и/или старый кнопочный сценарий.",
    )


# ─── /internal/parse_template ──────────────────────────────────────


class ParseTemplateIn(BaseModel):
    xlsx_path: str                      # абсолютный путь к xlsx-файлу на сервере
    cabinet: str | None = None          # имя кабинета или "default"
    save_as: str | None = None          # имя для сохранения (без расширения)


class ParseTemplateOut(BaseModel):
    ok: bool
    saved_to: str | None = None
    marketplace: str | None = None
    sheet_name: str | None = None
    n_fields: int = 0
    n_required: int = 0
    n_with_dropdown: int = 0
    category_id: int | None = None
    type_id: int | None = None
    parse_warnings: list[str] = []
    error: str | None = None


@router.post("/parse_template", response_model=ParseTemplateOut)
async def parse_template_endpoint(
    req: ParseTemplateIn,
    request: Request,
    x_internal_token: str | None = Header(default=None, alias="X-Internal-Token"),
):
    """Парсит xlsx-шаблон Ozon/WB через app.excel.parser, сохраняет JSON-структуру
    в ~/cz-backend/templates/<cabinet>/<marketplace>_<save_as>.json.

    Используется скиллом гнома `parse_template` как первый шаг Excel-флоу.
    """
    _check_token(x_internal_token)
    import dataclasses
    import json
    from pathlib import Path

    from .excel.parser import parse_template as _parse

    p = Path(req.xlsx_path).expanduser()
    if not p.exists():
        return ParseTemplateOut(ok=False, error=f"файл не найден: {req.xlsx_path}")
    if p.suffix.lower() != ".xlsx":
        return ParseTemplateOut(ok=False, error=f"не xlsx: {p.suffix}")

    try:
        spec = _parse(p)
    except Exception as e:
        logger.exception("parse_template fail %s", req.xlsx_path)
        return ParseTemplateOut(ok=False, error=f"парсинг упал: {str(e)[:300]}")

    cabinet = (req.cabinet or "default").strip() or "default"
    save_name = (req.save_as or p.stem).strip() or p.stem
    out_dir = Path.home() / "cz-backend" / "templates" / cabinet
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"{spec.marketplace}_{save_name}.json"

    try:
        out_path.write_text(
            json.dumps(dataclasses.asdict(spec), ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
    except Exception as e:
        return ParseTemplateOut(ok=False, error=f"сохранение упало: {str(e)[:200]}")

    n_required = sum(1 for f in spec.fields if f.required)
    n_with_dd = sum(1 for f in spec.fields if f.dropdown)

    # Кладём оригинал xlsx и meta.json в template_cache, чтобы потом
    # find_template(cabinet, mp, category_id) находил их при сборке партии.
    # Cache переживает рестарт сервиса.
    try:
        from .template_cache import save_template
        if spec.category_id:
            save_template(
                cabinet=cabinet,
                source_xlsx=p,
                parsed_meta_json=out_path,
                marketplace=spec.marketplace,
                category_id=spec.category_id,
            )
    except Exception as e:
        logger.warning("template_cache: save failed for %s: %s",
                       req.xlsx_path, str(e)[:200])

    return ParseTemplateOut(
        ok=True,
        saved_to=str(out_path),
        marketplace=spec.marketplace,
        sheet_name=spec.sheet_name,
        n_fields=len(spec.fields),
        n_required=n_required,
        n_with_dropdown=n_with_dd,
        category_id=spec.category_id,
        type_id=spec.type_id,
        parse_warnings=spec.parse_warnings or [],
    )


# ─── /internal/fill_excel_batch ────────────────────────────────────


class ProductInput(BaseModel):
    sku: str
    name: str
    brand: str = ""
    weight_g: int | None = None              # вес одиночки в граммах
    dims: dict[str, float] | None = None     # {l,w,h} см
    images: dict[str, str] = {}              # tag → URL


class PendingQuestion(BaseModel):
    field_id: str                            # уникальный для answers (sku::field)
    field_name: str
    sku: str
    title: str
    options: list[dict] = []                 # {label, score?, id?, detail?}
    allow_freetext: bool = True


class FillExcelBatchIn(BaseModel):
    template_json_path: str
    products: list[ProductInput]
    cabinet: str | None = None
    answers: dict[str, str] = {}             # field_id → answer
    output_filename: str | None = None       # имя сохраняемого xlsx


class FillExcelBatchOut(BaseModel):
    ok: bool
    state: str                               # "filled" | "pending" | "error"
    xlsx_path: str | None = None
    pending: list[PendingQuestion] = []
    skus_total: int = 0
    skus_filled: int = 0
    error: str | None = None


_FIELD_NAME_MAPPINGS = {
    # Ключевое слово в имени поля → ключ-резолвер (lower-case)
    "артикул": "sku",
    "vendorcode": "sku",
    "название товара": "title",
    "наименование": "title_wb_full",
    "бренд": "brand",
    "описание": "description",
    "аннотация": "description",
    "категория продавца": "category_path",
    "вес товара": "weight_unit_g",
    "вес в упаковке": "weight_packed_g",
    "вес брутто": "weight_packed_g",
    "длина": "dim_l",
    "ширина": "dim_w",
    "высота": "dim_h",
    "глубина": "dim_l",
    "ндс": "vat",
    "цена, руб": "price",
    "цена до скидки": "price_old",
    "штрих": "barcode_skip",
    "артикул wb": "wb_article_skip",
    "группа": "group_name",
    "фото": "image_main",
}


def _resolve_field_key(field_name: str) -> str | None:
    """По имени поля шаблона определяет универсальный ключ-резолвер.
    Возвращает None если поле «нестандартное» (атрибут категории).
    """
    nm = (field_name or "").lower()
    for kw, key in _FIELD_NAME_MAPPINGS.items():
        if kw in nm:
            return key
    return None


def _det_value_for_sku(
    field_key: str,
    sku_row: dict[str, Any],
    parsed: dict[str, str],
    product: ProductInput,
    marketplace: str,
    category_path: str,
) -> str | None:
    """Детерминированное значение для известного field_key. None = пропустить."""
    from .normalize import format_ozon_title, format_wb_full_title, format_wb_short_title
    from .rules import nds_value

    qty = sku_row.get("qty", 1)
    dims = sku_row.get("dims") or {}
    if field_key == "sku":
        return sku_row["sku"]
    if field_key == "title":
        if marketplace == "ozon":
            return format_ozon_title(parsed, qty=qty)
        return format_wb_full_title(parsed, qty=qty)
    if field_key == "title_wb_full":
        if marketplace == "wb":
            return format_wb_full_title(parsed, qty=qty)
        return format_ozon_title(parsed, qty=qty)
    if field_key == "title_wb_short":
        return format_wb_short_title(parsed)
    if field_key == "brand":
        return product.brand or parsed.get("brand", "")
    if field_key == "description":
        # placeholder — реальная аннотация генерится LLM (этап 3 батчевый),
        # здесь пишем хотя бы имя товара чтобы поле не было пустым.
        return product.name
    if field_key == "category_path":
        return category_path
    if field_key == "weight_unit_g":
        v = sku_row.get("weight_unit_g")
        return str(v) if v else None
    if field_key == "weight_packed_g":
        if marketplace == "wb":
            v = sku_row.get("weight_wb_kg")
            return str(v) if v else None
        v = sku_row.get("weight_packed_g")
        return str(v) if v else None
    if field_key == "dim_l":
        v = dims.get("l")
        return str(v) if v else None
    if field_key == "dim_w":
        v = dims.get("w")
        return str(v) if v else None
    if field_key == "dim_h":
        v = dims.get("h")
        return str(v) if v else None
    if field_key == "vat":
        return str(nds_value())  # 22
    if field_key == "price":
        return "0"
    if field_key == "price_old":
        return ""
    if field_key in ("barcode_skip", "wb_article_skip"):
        return ""
    if field_key == "group_name":
        # WB-группа: bren_subjectId или просто бренд
        brand = product.brand or parsed.get("brand", "")
        return brand or product.sku
    if field_key == "image_main":
        return product.images.get("main", "")
    return None


async def _resolve_category_path(spec, deps: Deps, cabinet_name: str | None = None) -> str:
    """Возвращает читаемый путь категории для поля «Категория продавца».

    Приоритет:
      1. Ozon + есть category_id → ищем в дереве категорий (с fallback по кабинетам).
      2. Имя файла шаблона без расширения (например «Стиральные порошки»).
      3. Пусто.
    """
    from pathlib import Path

    if spec.marketplace == "ozon" and spec.category_id:
        ozon_client, _ = _pick_ozon_client(cabinet_name, deps)
        if ozon_client is not None:
            try:
                tree = await ozon_client.category_tree()
                path = _find_ozon_path_by_id(tree, int(spec.category_id))
                if path:
                    return path
            except Exception as e:
                logger.warning("ozon category tree resolve failed: %s", e)

    raw = getattr(spec, "raw_path", "") or ""
    if raw:
        return Path(raw).stem.strip()
    return ""


def _find_ozon_path_by_id(tree: list[dict], target_id: int, prefix: str = "") -> str:
    """Рекурсивный поиск Ozon-категории по description_category_id."""
    for n in tree:
        name = n.get("category_name") or n.get("type_name") or ""
        cur = f"{prefix} / {name}" if prefix else name
        cat_id = n.get("description_category_id")
        if cat_id and int(cat_id) == target_id:
            return cur
        children = n.get("children") or n.get("types") or []
        if children:
            found = _find_ozon_path_by_id(children, target_id, cur)
            if found:
                return found
    return ""


@router.post("/fill_excel_batch", response_model=FillExcelBatchOut)
async def fill_excel_batch_endpoint(
    req: FillExcelBatchIn,
    request: Request,
    x_internal_token: str | None = Header(default=None, alias="X-Internal-Token"),
):
    """Заполняет xlsx-шаблон по 3 SKU на товар через правила §5.2.

    Двухфазный flow:
      1. Первый вызов (answers пуст): прогоняет детерминированно, на required-полях
         с dropdown'ом возвращает pending-вопросы.
      2. Второй вызов (answers заполнен): применяет ответы, пишет xlsx, возвращает путь.
    """
    _check_token(x_internal_token)
    import dataclasses
    import json
    from pathlib import Path

    from openpyxl import load_workbook

    from .excel.parser import TemplateField, TemplateSpec
    from .normalize import parse_input_line
    from .rules import expand_to_3_skus

    # 1. Восстанавливаем TemplateSpec из JSON
    json_p = Path(req.template_json_path).expanduser()
    if not json_p.exists():
        return FillExcelBatchOut(ok=False, state="error",
                                 error=f"template json не найден: {json_p}")
    try:
        spec_dict = json.loads(json_p.read_text(encoding="utf-8"))
        fields = [TemplateField(**f) for f in spec_dict.get("fields") or []]
        spec_dict["fields"] = fields
        spec = TemplateSpec(**spec_dict)
    except Exception as e:
        return FillExcelBatchOut(ok=False, state="error",
                                 error=f"json parse failed: {str(e)[:200]}")

    if not spec.raw_path or not Path(spec.raw_path).exists():
        return FillExcelBatchOut(
            ok=False, state="error",
            error=f"оригинал xlsx не найден: {spec.raw_path}",
        )

    # 2. Разворачиваем продукты на 3 SKU + парсим имена
    sku_rows: list[dict[str, Any]] = []   # развёрнутые SKU с их продуктом
    for product in req.products:
        parsed = parse_input_line(product.name, brand_hint=product.brand)
        rows = expand_to_3_skus({
            "sku": product.sku,
            "name": product.name,
            "weight": product.weight_g or 0,
            "dims": product.dims or {"l": 15, "w": 10, "h": 5},
        }, dims_from_internet=True)
        for r in rows:
            r["_product"] = product
            r["_parsed"] = parsed
            sku_rows.append(r)

    if not sku_rows:
        return FillExcelBatchOut(ok=False, state="error", error="продуктов нет")

    # 3. Резолвим category_path по category_id (Ozon) или из имени файла шаблона
    deps: Deps = request.app.state.deps
    category_path = await _resolve_category_path(spec, deps, req.cabinet)

    # 3b. Подтягиваем cached answers per-product (по бренду+имени)
    from .decision_cache import append_cache, read_cached_answers
    cabinet_norm = (req.cabinet or "default").strip() or "default"
    cached_per_sku: dict[str, dict[str, str]] = {}
    for product in req.products:
        cached = read_cached_answers(
            cabinet_norm, spec.marketplace, product.brand, product.name,
        )
        # Распространяем cached на все 3 SKU этого product
        for sku_row in sku_rows:
            if sku_row["_product"].sku == product.sku:
                cached_per_sku[sku_row["sku"]] = cached
    pending: list[PendingQuestion] = []
    filled_rows: list[dict[int, str]] = []  # для каждого SKU: column → value

    for sku_row in sku_rows:
        product: ProductInput = sku_row["_product"]
        parsed = sku_row["_parsed"]
        row_values: dict[int, str] = {}
        cached = cached_per_sku.get(sku_row["sku"], {})
        for field in spec.fields:
            field_key = _resolve_field_key(field.name)
            value: str | None = None
            answer_key = f"{sku_row['sku']}::{field.name}"

            # 3a. Сначала проверяем явный answer от юзера
            if answer_key in req.answers:
                ans = req.answers[answer_key].strip()
                if ans:
                    value = ans

            # 3b. Иначе — детерминированно
            if value is None and field_key:
                value = _det_value_for_sku(
                    field_key, sku_row, parsed, product,
                    spec.marketplace, category_path,
                )

            # 3b'. Иначе — из persistent cache decisions.jsonl
            if (value is None or value == "") and field.name in cached:
                value = cached[field.name]

            # 3c. Если поле required и до сих пор пусто — pending
            if (value is None or value == "") and field.required:
                # Подготовим options для ask_user_choice
                options: list[dict] = []
                if field.dropdown:
                    for v in field.dropdown[:5]:
                        options.append({"label": v})
                pending.append(PendingQuestion(
                    field_id=answer_key,
                    field_name=field.name,
                    sku=sku_row["sku"],
                    title=f"{sku_row['sku']} — поле «{field.name}»: какое значение?",
                    options=options,
                    allow_freetext=True,
                ))
                continue

            # 3d. Не required и не нашли — пропускаем (xlsx останется пустым)
            if value is None or value == "":
                continue

            row_values[field.column] = value
        filled_rows.append(row_values)

    # 4. Если есть pending — возвращаем (без записи xlsx)
    if pending:
        return FillExcelBatchOut(
            ok=True, state="pending",
            pending=pending,
            skus_total=len(sku_rows),
            skus_filled=0,
        )

    # 5. Все required заполнены — пишем xlsx
    try:
        wb = load_workbook(spec.raw_path, data_only=False)
        ws = wb[spec.sheet_name] if spec.sheet_name in wb.sheetnames else wb.active

        for i, row_values in enumerate(filled_rows):
            target_row = spec.data_start_row + i
            for col, val in row_values.items():
                ws.cell(row=target_row, column=col, value=val)

        cabinet = (req.cabinet or "default").strip() or "default"
        out_dir = Path.home() / "cz-backend" / "filled" / cabinet
        out_dir.mkdir(parents=True, exist_ok=True)
        fname = req.output_filename or f"{Path(spec.raw_path).stem}_filled.xlsx"
        out_path = out_dir / fname
        wb.save(out_path)
    except Exception as e:
        logger.exception("fill_excel_batch save failed")
        return FillExcelBatchOut(ok=False, state="error",
                                 error=f"save failed: {str(e)[:200]}")

    # 6. Записываем в persistent cache те answers что юзер дал в этой партии
    #    (только настоящие user-answers, не наши авто-fill).
    for product in req.products:
        # собираем по всем 3 SKU этого продукта
        merged: dict[str, str] = {}
        for sku_row in sku_rows:
            if sku_row["_product"].sku != product.sku:
                continue
            sku = sku_row["sku"]
            prefix = f"{sku}::"
            for k, v in (req.answers or {}).items():
                if k.startswith(prefix) and v:
                    field_name = k[len(prefix):]
                    merged[field_name] = v
        if merged:
            try:
                append_cache(cabinet_norm, spec.marketplace,
                             product.brand, product.name, merged)
            except Exception as e:
                logger.warning("cache append fail %s: %s", product.sku, e)

    return FillExcelBatchOut(
        ok=True, state="filled",
        xlsx_path=str(out_path),
        skus_total=len(sku_rows),
        skus_filled=len(sku_rows),
    )


# ─── /internal/load_wb_dropdowns ───────────────────────────────────


class LoadWbDropdownsIn(BaseModel):
    template_json_path: str             # путь к JSON от parse_template
    cabinet: str | None = None          # имя кабинета (если есть несколько WB)


class LoadWbDropdownsOut(BaseModel):
    ok: bool
    fields_updated: int = 0
    fields_with_values: int = 0
    error: str | None = None


@router.post("/load_wb_dropdowns", response_model=LoadWbDropdownsOut)
async def load_wb_dropdowns_endpoint(
    req: LoadWbDropdownsIn,
    request: Request,
    x_internal_token: str | None = Header(default=None, alias="X-Internal-Token"),
):
    """Подтягивает значения dropdown'ов для WB-полей с charcType=4/5.

    WB-шаблон в xlsx не содержит сами значения (в отличие от Ozon — там они
    на листе validation). Только charcID и charcType через defined_names.
    Этот endpoint идёт по полям где wb_charc_type in (4,5), дёргает
    WBClient.subject_charcs/directory_values, складывает в TemplateField.dropdown
    и перезаписывает JSON.
    """
    _check_token(x_internal_token)
    import dataclasses
    import json
    from pathlib import Path

    from .excel.parser import TemplateField, TemplateSpec
    from .wb import WBError

    json_p = Path(req.template_json_path).expanduser()
    if not json_p.exists():
        return LoadWbDropdownsOut(ok=False, error=f"json не найден: {json_p}")

    try:
        spec_dict = json.loads(json_p.read_text(encoding="utf-8"))
        fields = [TemplateField(**f) for f in spec_dict.get("fields") or []]
        spec_dict["fields"] = fields
        spec = TemplateSpec(**spec_dict)
    except Exception as e:
        return LoadWbDropdownsOut(ok=False, error=f"json parse failed: {str(e)[:200]}")

    if spec.marketplace != "wb":
        return LoadWbDropdownsOut(ok=False, error=f"не WB-шаблон: {spec.marketplace}")

    deps: Deps = request.app.state.deps
    wb_client, used_label = _pick_wb_client(req.cabinet, deps)
    if wb_client is None:
        return LoadWbDropdownsOut(
            ok=False,
            error="ни в одном кабинете не настроен WB_TOKEN",
        )

    # Идём по dictionary-полям (charcType 4/5). Дёргаем directory_values
    # для каждого charcID — у некоторых WB-эндпоинтов директория именная,
    # тут используем charcID как имя (WB API принимает {name} в URL).
    n_updated = 0
    n_with_values = 0
    for f in fields:
        if f.wb_charc_type not in (4, 5):
            continue
        if not f.wb_charc_id:
            continue
        try:
            vals = await wb_client.directory_values(str(f.wb_charc_id))
        except WBError as e:
            logger.warning("wb directory %s failed: %s", f.wb_charc_id, e)
            continue
        except Exception as e:
            logger.warning("wb directory %s exception: %s", f.wb_charc_id, e)
            continue
        names = [v.get("name") for v in vals if v.get("name")]
        if names:
            f.dropdown = names
            n_with_values += 1
        n_updated += 1

    # Перезаписываем JSON с обновлёнными dropdown'ами
    try:
        spec.fields = fields  # type: ignore[assignment]
        json_p.write_text(
            json.dumps(dataclasses.asdict(spec), ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
    except Exception as e:
        return LoadWbDropdownsOut(ok=False, error=f"json save failed: {str(e)[:200]}")

    return LoadWbDropdownsOut(
        ok=True,
        fields_updated=n_updated,
        fields_with_values=n_with_values,
    )


# ─── /internal/deliver_excel ───────────────────────────────────────


class DeliverExcelIn(BaseModel):
    xlsx_path: str
    chat_id: int
    caption: str | None = None
    filename: str | None = None      # переопределить имя при отправке


class DeliverExcelOut(BaseModel):
    ok: bool
    sent_filename: str | None = None
    size_bytes: int = 0
    error: str | None = None


@router.post("/deliver_excel", response_model=DeliverExcelOut)
async def deliver_excel_endpoint(
    req: DeliverExcelIn,
    request: Request,
    x_internal_token: str | None = Header(default=None, alias="X-Internal-Token"),
):
    """Отправляет готовый xlsx юзеру в Telegram через TelegramClient.send_document."""
    _check_token(x_internal_token)
    from pathlib import Path

    p = Path(req.xlsx_path).expanduser()
    if not p.exists():
        return DeliverExcelOut(ok=False, error=f"файл не найден: {req.xlsx_path}")
    if p.suffix.lower() != ".xlsx":
        return DeliverExcelOut(ok=False, error=f"не xlsx: {p.suffix}")

    try:
        content = p.read_bytes()
    except Exception as e:
        return DeliverExcelOut(ok=False, error=f"чтение упало: {str(e)[:200]}")

    fname = req.filename or p.name
    deps: Deps = request.app.state.deps
    try:
        await deps.tg.send_document(
            chat_id=req.chat_id,
            content=content,
            filename=fname,
            caption=req.caption or f"📋 *{fname}*",
            parse_mode="Markdown",
        )
    except Exception as e:
        logger.exception("deliver_excel TG send failed")
        return DeliverExcelOut(ok=False, error=f"telegram: {str(e)[:200]}")

    return DeliverExcelOut(ok=True, sent_filename=fname, size_bytes=len(content))


# ─── /internal/run_full_batch ──────────────────────────────────────


class RunBatchProductIn(BaseModel):
    sku: str
    name: str
    brand: str | None = None
    src_url: str  # фото уже в S3 (bridge или гном залил)
    weight_g: int | None = None
    dims: dict[str, float] | None = None  # {"l":..,"w":..,"h":..}


class RunBatchIn(BaseModel):
    chat_id: int
    products: list[RunBatchProductIn]
    cabinet_names: list[str] | None = None  # None = default; ["all"] не поддерживаем
    dry_run: bool | None = None  # None = берём из settings


class RunBatchOut(BaseModel):
    ok: bool
    batch_id: str
    queued: bool
    products_count: int
    cabinets: list[str]
    dry_run: bool
    note: str = ""


@router.post("/run_full_batch", response_model=RunBatchOut)
async def run_full_batch_endpoint(
    req: RunBatchIn,
    request: Request,
    x_internal_token: str | None = Header(default=None, alias="X-Internal-Token"),
):
    """Запускает полный pipeline партии в фоне. Возвращается сразу с batch_id.

    Прогресс уходит юзеру в чат через `deps.tg.send` из `pipeline.run_batch`
    (типовые сообщения: «📥 фото в S3», «🎨 identity», тайминги этапов,
    финальный отчёт). Гному ничего ждать не надо.
    """
    _check_token(x_internal_token)
    if not req.products:
        raise HTTPException(status_code=400, detail="products пустой")
    if len(req.products) > 10:
        raise HTTPException(status_code=400,
                            detail="максимум 10 товаров на партию (см. ТЗ §1)")

    deps: Deps = request.app.state.deps
    from .models import ProductIn, RunRequest
    from .pipeline import run_batch
    import asyncio

    products = [
        ProductIn(
            idx=i,
            sku=p.sku,
            name=p.name,
            brand=p.brand,
            tg_file_id="",      # фото уже в S3, не качаем из TG
            src_url=p.src_url,
            weight_g=p.weight_g,
            dims=p.dims,
        )
        for i, p in enumerate(req.products)
    ]
    batch_id = f"gnome-{int(time.time())}-{uuid.uuid4().hex[:6]}"
    # Если LLM не передал cabinet_names — оставляем None, pipeline сам
    # разрешит default через settings.default_cabinet_name внутри run_batch.
    cabinet_names = req.cabinet_names or None
    rr = RunRequest(
        batch_id=batch_id,
        chat_id=req.chat_id,
        products=products,
        cabinet_names=cabinet_names,
    )

    # DRY_RUN override на уровне текущего process'а: если запросили — временно
    # включаем, восстанавливаем после старта таски (пока остаются в памяти,
    # это потокоопасно если несколько батчей одновременно — пусть пока будет
    # глобальный settings.DRY_RUN, override игнорим).
    asyncio.create_task(run_batch(rr, deps))

    cabs_used = cabinet_names or ["default"]
    return RunBatchOut(
        ok=True,
        batch_id=batch_id,
        queued=True,
        products_count=len(products),
        cabinets=cabs_used,
        dry_run=settings.DRY_RUN,
        note=("Партия запущена в фоне. Прогресс приходит в чат отдельными "
              "сообщениями от пайплайна."),
    )


# ─── /internal/make_batch_zip — главный «гениальный» endpoint ─────────
# Партия → фото → категории → проверка кеша шаблонов → заполнение →
# ZIP с photos/+ozon/+wb/+README.txt → отправка юзеру в TG. БЕЗ API
# заливки на маркетплейсы — юзер сам грузит файл в свой кабинет.


class MakeBatchZipIn(BaseModel):
    chat_id: int
    products: list[RunBatchProductIn]   # переиспользуем модель из run_full_batch
    cabinet: str | None = None


class MissingTemplateInfo(BaseModel):
    mp: str                              # ozon | wb
    category_id: int
    category_path: str = ""


class MakeBatchZipOut(BaseModel):
    ok: bool
    batch_id: str
    need_templates: list[MissingTemplateInfo] = []
    zip_sent: bool = False
    n_products: int = 0
    n_skus: int = 0
    n_xlsx: int = 0
    error: str | None = None


async def _run_zip_pipeline(req: MakeBatchZipIn, deps: Deps, batch_id: str) -> None:
    """Фоновая таска: фото → категории → шаблоны → заполнение → ZIP → TG.

    Любая ошибка ловится и отчитывается юзеру в чат, чтобы фоновая корутина
    не падала молча.
    """
    from pathlib import Path

    from .models import ProductIn, ProductState
    from .pipeline import (
        process_product_images,
        match_category,
        _flatten_tree,
    )
    from .template_cache import find_template
    from .zip_builder import build_batch_zip

    chat_id = req.chat_id
    cabinet = req.cabinet or settings.default_cabinet_name

    try:
        await deps.tg.send(
            chat_id,
            f"🟦 *Партия* `{batch_id}` собирается ({len(req.products)} товаров)",
        )
    except Exception:
        pass

    # 1. ProductState
    products_in = [
        ProductIn(
            idx=i, sku=p.sku, name=p.name, brand=p.brand,
            tg_file_id="", src_url=p.src_url,
            weight_g=p.weight_g, dims=p.dims,
        )
        for i, p in enumerate(req.products)
    ]
    states = [ProductState.from_in(p) for p in products_in]

    # 2. Этап 1 — фото на каждый товар (4 шт в 3:4)
    sem = asyncio.Semaphore(settings.MAX_PARALLEL_PRODUCTS)
    try:
        await asyncio.gather(*[
            process_product_images(s, batch_id, chat_id, deps, sem) for s in states
        ])
    except Exception as e:
        logger.exception("zip batch: фото-этап упал")
        try:
            await deps.tg.send(chat_id, f"❌ Этап фото упал: {str(e)[:200]}")
        except Exception:
            pass
        return

    # 3. Этап 2 — категории
    try:
        await deps.tg.send(chat_id, "📂 Подбираю категории Ozon и WB…")
    except Exception:
        pass

    ozon_leaves: list[dict] = []
    wb_leaves: list[dict] = []
    try:
        ozon_tree = await deps.ozon.category_tree()
        ozon_leaves = _flatten_tree(ozon_tree, is_ozon=True)
    except Exception as e:
        logger.warning("ozon.category_tree fail: %s", e)
    try:
        wb_tree = await deps.wb.subjects_tree()
        wb_leaves = _flatten_tree(wb_tree, is_ozon=False)
    except Exception as e:
        logger.warning("wb.subjects_tree fail: %s", e)

    if ozon_leaves or wb_leaves:
        await asyncio.gather(*[
            match_category(s, ozon_leaves or [], wb_leaves or [], deps)
            for s in states
        ])

    # 4. Lookup шаблонов в кеше
    needed: dict[tuple[str, int], str] = {}   # (mp, cat_id) → category_path
    for s in states:
        if s.ozon_category and s.ozon_category.id:
            cid = int(s.ozon_category.id)
            needed.setdefault(("ozon", cid), s.ozon_category.path or "")
        if s.wb_subject and s.wb_subject.id:
            sid = int(s.wb_subject.id)
            needed.setdefault(("wb", sid), s.wb_subject.path or "")

    found: dict[tuple[str, int], object] = {}
    missing: list[MissingTemplateInfo] = []
    for (mp, cid), path in needed.items():
        t = find_template(cabinet, mp, cid)
        if t is not None:
            found[(mp, cid)] = t
        else:
            missing.append(MissingTemplateInfo(mp=mp, category_id=cid, category_path=path))

    if missing:
        lines = ["📋 *Нужны пустые xlsx-шаблоны от тебя:*", ""]
        for m in missing:
            lines.append(f"• *{m.mp.upper()}*: {m.category_path or m.category_id} (id={m.category_id})")
        lines.append("")
        lines.append("Скачай в кабинете МП пустой шаблон для этой категории и "
                     "кинь файлом сюда — я запомню навсегда. После этого "
                     "напиши «собрать ещё раз» — я доделаю партию.")
        try:
            await deps.tg.send(chat_id, "\n".join(lines), parse_mode="Markdown")
        except Exception:
            pass
        return

    # 5. Заполнение шаблонов — группируем по (mp, cat_id), вызываем
    # /internal/fill_excel_batch локально через httpx (один self-loop).
    try:
        await deps.tg.send(chat_id, "✏️ Заполняю шаблоны по правилам ТЗ…")
    except Exception:
        pass

    import httpx as _hx

    grouped_products: dict[tuple[str, int], list[dict]] = {}
    for s in states:
        # переводим state → ProductInput для fill_excel_batch
        prod = {
            "sku": s.sku,
            "name": s.name,
            "brand": s.brand or "",
            "weight_g": None,
            "dims": None,
            "images": s.images or {},
        }
        if s.ozon_category and s.ozon_category.id:
            grouped_products.setdefault(("ozon", int(s.ozon_category.id)), []).append(prod)
        if s.wb_subject and s.wb_subject.id:
            grouped_products.setdefault(("wb", int(s.wb_subject.id)), []).append(prod)

    xlsx_paths: dict[str, Path] = {}
    n_skus = 0

    auth_header = {"Content-Type": "application/json"}
    if settings.INTERNAL_TOKEN:
        auth_header["X-Internal-Token"] = settings.INTERNAL_TOKEN

    for (mp, cid), prods in grouped_products.items():
        t = found[(mp, cid)]
        body = {
            "template_json_path": t.json_path,
            "products": prods,
            "cabinet": cabinet or "default",
            "answers": {},
            "output_filename": f"{batch_id}_{mp}_{cid}.xlsx",
        }
        try:
            async with _hx.AsyncClient(timeout=120.0) as http:
                r = await http.post(
                    "http://127.0.0.1:8000/internal/fill_excel_batch",
                    headers=auth_header, json=body,
                )
            if r.status_code >= 400:
                raise RuntimeError(f"fill_excel_batch HTTP {r.status_code}: {r.text[:200]}")
            data = r.json()
        except Exception as e:
            logger.exception("fill_excel_batch fail %s/%s", mp, cid)
            try:
                await deps.tg.send(
                    chat_id, f"⚠️ Заполнение {mp} cat={cid} упало: {str(e)[:200]}",
                )
            except Exception:
                pass
            continue

        state = data.get("state")
        if state == "filled":
            xlsx_path = data.get("xlsx_path")
            if xlsx_path and Path(xlsx_path).exists():
                arc = f"{mp}/{batch_id}_{mp}_{cid}.xlsx"
                xlsx_paths[arc] = Path(xlsx_path)
                n_skus += int(data.get("skus_filled") or 0)
        elif state == "pending":
            # Есть незаполненные required-поля. Кладём пустой шаблон + предупреждение.
            n_pending = len(data.get("pending") or [])
            try:
                await deps.tg.send(
                    chat_id,
                    f"⚠️ {mp.upper()} cat={cid}: {n_pending} полей "
                    "не заполнены автоматически — допиши их в Excel перед загрузкой.",
                )
            except Exception:
                pass
            # Шаблон-исходник тоже кладём в ZIP (без заполненных строк)
            arc = f"{mp}/EMPTY_{batch_id}_{mp}_{cid}.xlsx"
            xlsx_paths[arc] = Path(t.xlsx_path)
        else:
            try:
                await deps.tg.send(
                    chat_id, f"❌ {mp} cat={cid}: {data.get('error') or 'unknown error'}",
                )
            except Exception:
                pass

    if not xlsx_paths:
        try:
            await deps.tg.send(chat_id, "❌ Ни один шаблон не собран. Партия не уйдёт в ZIP.")
        except Exception:
            pass
        return

    # 6. ZIP-упаковка + отправка юзеру
    try:
        await deps.tg.send(chat_id, "📦 Собираю архив…")
    except Exception:
        pass

    photo_urls = {s.sku: dict(s.images) for s in states if s.images}
    # для скачивания фоток из S3 используем общий httpx-клиент cz-backend;
    # если его нет — создаём временный
    if deps.http is not None:
        zip_http = deps.http
        own_zip_http = False
    else:
        zip_http = _hx.AsyncClient(timeout=60.0)
        own_zip_http = True
    try:
        try:
            zip_bytes = await build_batch_zip(
                batch_id=batch_id,
                photo_urls=photo_urls,
                xlsx_paths=xlsx_paths,
                http=zip_http,
                n_products=len(states),
                n_skus=n_skus or len(states) * 3,
            )
        finally:
            if own_zip_http:
                await zip_http.aclose()
    except Exception as e:
        logger.exception("zip build fail")
        try:
            await deps.tg.send(chat_id, f"❌ Сборка ZIP упала: {str(e)[:200]}")
        except Exception:
            pass
        return

    try:
        caption = (
            f"✅ Партия `{batch_id}` готова\n\n"
            f"📦 {len(states)} товаров • {len(xlsx_paths)} xlsx-файл(а)\n"
            f"Распакуй и грузи: ozon/* в кабинет Ozon, wb/* в кабинет WB.\n"
            f"Фотки уже зашиты в xlsx как URL — отдельно грузить не надо."
        )
        await deps.tg.send_document(
            chat_id=chat_id,
            content=zip_bytes,
            filename=f"{batch_id}.zip",
            caption=caption,
            parse_mode="Markdown",
        )
    except Exception as e:
        logger.exception("send_document fail")
        try:
            await deps.tg.send(chat_id, f"❌ Отправка ZIP в TG упала: {str(e)[:200]}")
        except Exception:
            pass


@router.post("/make_batch_zip", response_model=MakeBatchZipOut)
async def make_batch_zip_endpoint(
    req: MakeBatchZipIn,
    request: Request,
    x_internal_token: str | None = Header(default=None, alias="X-Internal-Token"),
):
    """Главный endpoint новой архитектуры контент-завода.

    Гном вызывает этот endpoint, он стартует фоновую таску, юзер видит
    прогресс в TG отдельными сообщениями. На выходе — ZIP-документ с
    фотками и заполненными xlsx-шаблонами. БЕЗ API заливки на МП.
    """
    _check_token(x_internal_token)
    if not req.products:
        raise HTTPException(status_code=400, detail="products пустой")
    if len(req.products) > 10:
        raise HTTPException(status_code=400, detail="максимум 10 товаров на партию")

    deps: Deps = request.app.state.deps
    batch_id = f"zip-{int(time.time())}-{uuid.uuid4().hex[:6]}"

    asyncio.create_task(_run_zip_pipeline(req, deps, batch_id))
    return MakeBatchZipOut(
        ok=True,
        batch_id=batch_id,
        zip_sent=False,
        n_products=len(req.products),
        n_skus=0,
        n_xlsx=0,
    )
