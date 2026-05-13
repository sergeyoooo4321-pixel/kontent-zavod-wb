from openpyxl import load_workbook

from app.config import Settings
from app.excel_export import build_sku_variants, build_zip
from app.models import GeneratedImage, ProductInput, ProductResult


def test_build_sku_variants_titles_and_dims():
    settings = Settings(TG_BOT_TOKEN="x")
    product = ProductInput(
        photo_index=1,
        sku="59031",
        brand="Tide",
        name="Стиральный порошок 400 г",
        weight_g=400,
        length_cm=10,
        width_cm=6,
        height_cm=20,
    )
    variants = build_sku_variants(product, settings)
    assert [v.sku for v in variants] == ["59031", "59031x2", "59031x3"]
    assert variants[1].ozon_title.startswith("Набор 2 шт Tide")
    assert variants[1].width_cm == 12
    assert variants[2].width_cm == 18
    assert variants[2].weight_g == 1200


def test_build_zip_contains_workbooks(tmp_path):
    settings = Settings(TG_BOT_TOKEN="x")
    result = ProductResult(
        input=ProductInput(photo_index=1, sku="A1", brand="Brand", name="Product 100 г"),
        images=[
            GeneratedImage(role="main", url="https://s3/main.jpg", key="k/main", bytes_data=b"jpg"),
            GeneratedImage(role="pack2", url="https://s3/pack2.jpg", key="k/pack2", bytes_data=b"jpg"),
            GeneratedImage(role="pack3", url="https://s3/pack3.jpg", key="k/pack3", bytes_data=b"jpg"),
            GeneratedImage(role="extra", url="https://s3/extra.jpg", key="k/extra", bytes_data=b"jpg"),
        ],
    )
    zip_bytes = build_zip([result], settings)
    path = tmp_path / "pack.zip"
    path.write_bytes(zip_bytes)
    import zipfile

    with zipfile.ZipFile(path) as zf:
        names = set(zf.namelist())
        assert "ozon.xlsx" in names
        assert "wildberries.xlsx" in names
        assert "links.csv" in names
        ozon_path = tmp_path / "ozon.xlsx"
        ozon_path.write_bytes(zf.read("ozon.xlsx"))
    wb = load_workbook(ozon_path)
    ws = wb.active
    assert ws.max_row == 4
    assert ws["A2"].value == "A1"
    assert ws["A3"].value == "A1x2"

