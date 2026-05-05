"""Тесты парсера xlsx-шаблонов."""
from pathlib import Path

import pytest

from app.excel.parser import detect_format, parse_template
import openpyxl


FIXTURES = Path(__file__).parent / "fixtures" / "excel"


# ─── Ozon ────────────────────────────────────────────────────────


def test_parse_ozon_sample():
    spec = parse_template(FIXTURES / "sample_ozon.xlsx")
    assert spec.marketplace == "ozon"
    assert spec.sheet_name == "Шаблон"
    assert spec.header_row == 2
    assert spec.data_start_row == 5
    assert spec.category_id is not None
    assert len(spec.fields) >= 30


def test_parse_ozon_required_marker():
    spec = parse_template(FIXTURES / "sample_ozon.xlsx")
    by_name = {f.name: f for f in spec.fields}
    # «Артикул*» обязателен
    assert "Артикул" in by_name
    assert by_name["Артикул"].required is True
    # «Цена до скидки» — не обязательное
    if "Цена до скидки, руб." in by_name:
        assert by_name["Цена до скидки, руб."].required is False


def test_parse_ozon_has_dropdowns():
    spec = parse_template(FIXTURES / "sample_ozon.xlsx")
    with_dd = [f for f in spec.fields if f.dropdown]
    # «НДС, %», «Рассрочка», «Баллы за отзывы» — точно с dropdown
    assert len(with_dd) >= 3, "expected at least 3 dropdown fields"
    nds = next((f for f in spec.fields if "НДС" in f.name), None)
    assert nds is not None
    assert nds.dropdown is not None
    assert "22" in nds.dropdown


def test_parse_ozon_category_id_from_configs():
    spec = parse_template(FIXTURES / "sample_ozon.xlsx")
    assert spec.category_id is not None
    assert spec.category_id > 0


# ─── WB ──────────────────────────────────────────────────────────


def test_parse_wb_sample():
    spec = parse_template(FIXTURES / "sample_wb.xlsx")
    assert spec.marketplace == "wb"
    assert spec.sheet_name == "Товары"
    assert spec.header_row == 3
    assert spec.data_start_row == 5
    assert len(spec.fields) >= 20


def test_parse_wb_hard_required():
    spec = parse_template(FIXTURES / "sample_wb.xlsx")
    by_name = {f.name: f for f in spec.fields}
    for required_name in ("Артикул продавца", "Наименование",
                          "Категория продавца", "Бренд", "Описание"):
        assert required_name in by_name, f"missing field: {required_name}"
        assert by_name[required_name].required is True, \
            f"{required_name} should be required"


def test_parse_wb_readonly_fallback():
    """Шаблон с MultiCellRange-багом должен открываться через read_only fallback."""
    spec = parse_template(FIXTURES / "sample_wb_readonly.xlsx")
    assert spec.marketplace == "wb"
    assert len(spec.fields) > 0
    # warning должен сообщать о fallback
    assert any("read_only" in w for w in spec.parse_warnings)


def test_parse_wb_charc_metadata():
    """WB-парсер вытаскивает charcID/charcType из defined_names."""
    spec = parse_template(FIXTURES / "sample_wb.xlsx")
    # Должны быть и атрибуты товара и служебные поля
    charcs = [f for f in spec.fields if f.wb_charc_id]
    statics = [f for f in spec.fields if f.wb_static]
    assert len(charcs) >= 20, f"too few charcs: {len(charcs)}"
    assert len(statics) >= 5, f"too few statics: {len(statics)}"
    # Бренд должен иметь charc_id 14177446 (постоянный WB id)
    brand = next((f for f in spec.fields if f.name == "Бренд"), None)
    assert brand is not None
    assert brand.wb_charc_id == 14177446
    # «Вес товара без упаковки (г)» — charcType=4 (dictionary_single) через WB API
    weight = next((f for f in spec.fields
                   if "вес товара без упаковки" in f.name.lower()), None)
    assert weight is not None
    assert weight.wb_charc_type == 4
    assert weight.wb_charc_id == 89008


# ─── detect_format ───────────────────────────────────────────────


def test_detect_format_ozon():
    wb = openpyxl.load_workbook(FIXTURES / "sample_ozon.xlsx", data_only=False)
    assert detect_format(wb) == "ozon"


def test_detect_format_wb():
    wb = openpyxl.load_workbook(FIXTURES / "sample_wb.xlsx", data_only=False)
    assert detect_format(wb) == "wb"


# ─── unknown / error path ────────────────────────────────────────


def test_parse_unknown_raises(tmp_path: Path):
    """Левый xlsx без узнаваемых листов → ValueError."""
    import openpyxl
    p = tmp_path / "weird.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "RandomSheet"
    wb.save(p)
    with pytest.raises(ValueError, match="unknown template format"):
        parse_template(p)
